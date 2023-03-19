import requests
from bs4 import BeautifulSoup
import openpyxl

# Input pencarian user yang akan di cari
your_query = input("Enter your search query: ")

# Search URL berdasarkan query user
url = f"https://en.wikinews.org/w/index.php?title=Special:Search&limit=20&offset=0&ns0=1&search={your_query}"
response = requests.get(url)
soup = BeautifulSoup(response.content, "html.parser")

# Mencari artikel berita berdasarkan query user
articles = soup.find_all("li", attrs="mw-search-result")

# Membuat file excel baru
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "Search Results"

# Menginputkan data hasil scraping ke dalam worksheet excel
worksheet.cell(row=1, column=1, value="No.")
worksheet.cell(row=1, column=2, value="Title")
worksheet.cell(row=1, column=3, value="Date")
for i, article in enumerate(articles, start=2):
    title = article.find("div",attrs="mw-search-result-heading").text
    date = article.find("div", attrs="mw-search-result-data").text
    worksheet.cell(row=i, column=1, value=i-1)
    worksheet.cell(row=i, column=2, value=title)
    worksheet.cell(row=i, column=3, value=date)

# Menyimpan hasil input excel ke lokal
workbook.save(filename="Wikinews "+your_query+".xlsx")
print(f"Search results for '{your_query}' have been saved to Wikinews {your_query}.xlsx")
